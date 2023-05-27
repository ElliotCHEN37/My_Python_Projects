import random #“random”模塊用於產生隨機數
import string #“string”模塊用於快速調用字符與數字
import openpyxl

num = string.digits #導入數字
ABC = string.ascii_letters #導入所有ASCII大小寫字母
PSD = list(num + ABC) #密碼的樣式
random.shuffle(PSD) #使用“random”模塊打亂密碼順序
LEN = int(input("請問你需要多長的密碼: ")) #詢問密碼長度並將回答轉換為數字
PAS = "".join(PSD[:LEN]) #生成需要的長度的密碼

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
c1 = my_sheet.cell(row = 1, column = 1)
c1.value = "Password list"
c2 = my_sheet.cell(row = 2, column = 1)
c2.value = f"{PAS}"
my_wb.save("password.xlsx")

print(f"你的密碼已導出至 password.xlsx") #輸出密碼